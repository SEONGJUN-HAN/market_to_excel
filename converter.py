# -*- coding: utf-8 -*-
"""
오픈마켓 견적서 -> 엑셀 변환기 (브라우저/Pyodide 버전)

quote_to_excel.py 의 파싱·정리·엑셀 로직을 그대로 옮기되, 브라우저에서 돌도록 손본 것:
  - 입력이 파일 경로가 아니라 (파일명, bytes) 다. (업로드된 파일)
  - 11번가 PDF 는 데스크톱과 동일하게 pdfplumber 로 읽는다.
    (pdfplumber 는 pypdfium2(네이티브)를 의존성으로 달고 있어 그냥은 WASM 설치가 막히지만,
     pypdfium2 는 이미지 렌더링(display.py)에만 쓰이고 지연 import 라 텍스트 추출에는 안 쓰인다.
     그래서 micropip.install("pdfplumber", deps=False) 로 pypdfium2 를 건너뛰고 설치하면
     extract_text 결과가 데스크톱과 완전히 같다. pdfminer.six 직접 추출은 표 열이 흩어져 실패한다.)
  - Gemini 는 google-genai SDK 대신 브라우저 fetch(pyodide.http.pyfetch)로 REST 호출한다.
    (WASM 에는 스레드가 없으므로 ThreadPoolExecutor 대신 asyncio 로 동시 요청한다.)
  - tkinter/클립보드/파일저장은 없다. 엑셀은 bytes 로 돌려주고 다운로드는 JS 가 한다.

금액 규칙은 데스크톱판과 동일하다. '실제 결제 금액' 컬럼만 신뢰한다.
"""

import io
import re
import json
import asyncio
import datetime
from dataclasses import dataclass, field

from bs4 import BeautifulSoup


# ---------------------------------------------------------------- 공통

@dataclass
class Item:
    name: str = ""
    spec: str = ""
    qty: int = 0
    amount: int = 0
    note: str = ""
    mall: str = ""
    raw_name: str = ""
    raw_spec: str = ""

    def keep_raw(self):
        if not self.raw_name:
            self.raw_name = self.name
            self.raw_spec = self.spec
        return self


@dataclass
class Sheet:
    mall: str
    items: list = field(default_factory=list)
    shipping: int = 0
    stated_total: int = 0
    source: str = ""


def to_int(text):
    if not text:
        return 0
    digits = re.sub(r"[^\d-]", "", str(text))
    return int(digits) if digits and digits != "-" else 0


def allocate(prices, discount):
    """상품 단위 할인을 옵션 금액에 비례 배분한다. 끝수는 마지막 항목이 흡수한다."""
    if discount == 0 or not prices:
        return [0] * len(prices)
    base = sum(prices)
    if base == 0:
        return [0] * len(prices)
    out, acc = [], 0
    for p in prices[:-1]:
        d = round(discount * p / base)
        out.append(d)
        acc += d
    out.append(discount - acc)
    return out


def read_html(data):
    """업로드된 .xls(실은 HTML) 바이트를 문자열로. 지마켓/옥션은 cp949 인 경우가 많다."""
    for enc in ("utf-8", "cp949", "euc-kr"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", "replace")


def expand_grid(table):
    """rowspan/colspan 을 전개해 2차원 격자로 만든다. 각 칸은 <td> 태그 or None."""
    grid = {}
    for r, tr in enumerate(table.find_all("tr")):
        c = 0
        for cell in tr.find_all(["td", "th"]):
            while (r, c) in grid:
                c += 1
            rs = int(cell.get("rowspan", 1) or 1)
            cs = int(cell.get("colspan", 1) or 1)
            for dr in range(rs):
                for dc in range(cs):
                    grid[(r + dr, c + dc)] = cell
            c += cs
    if not grid:
        return []
    rows = max(k[0] for k in grid) + 1
    cols = max(k[1] for k in grid) + 1
    return [[grid.get((r, c)) for c in range(cols)] for r in range(rows)]


def txt(cell):
    return cell.get_text(" ", strip=True) if cell is not None else ""


# ---------------------------------------------------------------- 지마켓/옥션 (구식)

def parse_ebay_old(html, mall, source):
    soup = BeautifulSoup(html, "lxml")
    sheet = Sheet(mall=mall, source=source)

    table = None
    for t in soup.find_all("table"):
        head = t.find("thead")
        if head and "공급합계" in head.get_text():
            table = t
            break
    if table is None:
        raise ValueError("상품 표를 찾지 못했습니다.")

    body = table.find("tbody")
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        classes = tds[0].get("class") or []
        if "first" in classes and len(tds) >= 7:
            sheet.items.append(Item(
                name=txt(tds[1]),
                qty=to_int(txt(tds[3])),
                amount=to_int(txt(tds[6])),      # 공급합계 = 실결제
                note=txt(tds[2]),
                mall=mall,
            ))
        elif sheet.items:
            opt = txt(tds[0])
            if opt and opt != sheet.items[-1].name:
                sheet.items[-1].spec = opt

    for t in soup.find_all("table"):
        for tr in t.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            key = txt(cells[0])
            if key == "배송비":
                sheet.shipping = to_int(txt(cells[1]))
            elif "총 구매금액" in key:
                sheet.stated_total = to_int(txt(cells[1]))
    return sheet


# ---------------------------------------------------------------- 지마켓/옥션 (신식)

def parse_ebay_new(html, mall, source):
    soup = BeautifulSoup(html, "lxml")
    sheet = Sheet(mall=mall, source=source)

    table = None
    for t in soup.find_all("table"):
        if "공급합계" in t.get_text():
            table = t
            break
    if table is None:
        raise ValueError("상품 표를 찾지 못했습니다.")

    grid = expand_grid(table)

    hdr = None
    for row in grid:
        joined = [txt(c) for c in row]
        if "공급합계" in joined:
            hdr = joined
            break
    if hdr is None:
        raise ValueError("표 헤더를 찾지 못했습니다.")

    def col(*names):
        for i, h in enumerate(hdr):
            if any(n in h for n in names):
                return i
        return None

    c_qty = col("수량")
    c_price = col("공급가격", "공급가액")
    c_disc = col("할인금액")
    c_total = col("공급합계")
    c_seller = col("공급자명")
    c_name = col("상품명")

    groups = []
    seen_header = False
    for row in grid:
        cells = [txt(c) for c in row]
        if not seen_header:
            seen_header = cells == hdr
            continue
        if not any(cells):
            continue

        is_title = (
            c_name is not None and row[c_name] is not None
            and c_total is not None and row[c_total] is row[c_name]
        )
        if is_title:
            groups.append({"name": cells[c_name], "seller": "", "discount": 0,
                           "total": 0, "opts": []})
            continue

        if not groups:
            continue
        qty = to_int(cells[c_qty]) if c_qty is not None else 0
        price = to_int(cells[c_price]) if c_price is not None else 0
        if qty == 0 and price == 0:
            continue
        opt = cells[c_name] if c_name is not None else ""
        groups[-1]["opts"].append((opt, qty, price))
        groups[-1]["seller"] = cells[c_seller] if c_seller is not None else ""
        groups[-1]["discount"] = to_int(cells[c_disc]) if c_disc is not None else 0
        groups[-1]["total"] = to_int(cells[c_total]) if c_total is not None else 0

    for g in groups:
        prices = [p for _, _, p in g["opts"]]
        cuts = allocate(prices, g["discount"])
        for (opt, qty, price), cut in zip(g["opts"], cuts):
            spec = "" if opt in ("본품", g["name"]) else opt
            sheet.items.append(Item(
                name=g["name"], spec=spec, qty=qty,
                amount=price - cut, note=g["seller"], mall=mall,
            ))

    for li in soup.select("li.list-item"):
        label = li.select_one(".text__label")
        value = li.select_one(".text__value")
        if not label or not value:
            continue
        if "배송비" in label.get_text():
            sheet.shipping = to_int(value.get_text())
        elif "총 구매금액" in label.get_text():
            sheet.stated_total = to_int(value.get_text())
    return sheet


# ---------------------------------------------------------------- 11번가 PDF

ITEM_RE = re.compile(
    r"^(\d+)\s+(\d{6,})\s+(.+?)\s+(\d+)"
    r"\s+([\d,]+)원\s+([\d,]+)원\s+([\d,]+)원\s+([\d,]+)원$"
)
STOP_RE = re.compile(r"^(합계|배송비|총합계|https?://|No\b|수$|량$|판매가|\(할인전금액\))")


def is_cjk(ch):
    return bool(ch) and (
        "가" <= ch <= "힣"
        or "一" <= ch <= "鿿"
    )


def join_wrapped(parts):
    parts = [p for p in parts if p]
    if not parts:
        return ""
    s = parts[0]
    for nxt in parts[1:]:
        if s and is_cjk(s[-1]) and is_cjk(nxt[0]):
            s += nxt
        else:
            s += " " + nxt
    return s


def parse_11st(data, source):
    """텍스트 PDF. pdfplumber 로 페이지 텍스트를 뽑아 줄 단위로 읽는다.
    (표 오른쪽 테두리가 없어 격자 인식이 어긋나므로 줄 단위로 읽는 방식은 데스크톱과 같다.)"""
    import pdfplumber

    sheet = Sheet(mall="11번가", source=source)
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    lines = [ln.strip() for ln in text.split("\n")]

    pending = None

    def flush():
        nonlocal pending
        if not pending:
            return
        full = join_wrapped(pending["parts"])
        if " / " in full:
            name, spec = full.split(" / ", 1)
        else:
            name, spec = full, ""
        sheet.items.append(Item(
            name=name.strip(), spec=spec.strip(),
            qty=pending["qty"], amount=pending["amount"],
            note=pending["pno"], mall="11번가",
        ))
        pending = None

    for ln in lines:
        m = ITEM_RE.match(ln)
        if m:
            flush()
            pending = {
                "pno": m.group(2),
                "parts": [m.group(3).strip()],
                "qty": int(m.group(4)),
                "amount": to_int(m.group(8)),
            }
            continue
        if pending is not None:
            if STOP_RE.match(ln) or not ln:
                flush()
            else:
                pending["parts"].append(ln)
    flush()

    m = re.search(r"배송비\(선결제\)\s*([\d,]+)\s*원", text)
    if m:
        sheet.shipping += to_int(m.group(1))
    m = re.search(r"착불배송비\s*([\d,]+)\s*원", text)
    if m:
        sheet.shipping += to_int(m.group(1))
    m = re.search(r"총합계\s*([\d,]+)\s*원", text)
    if m:
        sheet.stated_total = to_int(m.group(1))
    return sheet


# ---------------------------------------------------------------- 몰 판별

MALL_NEW = "옥션"
MALL_OLD = "지마켓"
MALL_ORDER = ("지마켓", "옥션", "11번가")


def parse_file(name, data):
    """업로드된 (파일명, bytes) 하나를 Sheet 로 파싱한다."""
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if ext == "pdf":
        return parse_11st(data, name)

    html = read_html(data)
    if "11번가" in html or "십일번가" in html:
        raise ValueError("11번가 견적서는 PDF 로 받아주세요.")

    if "table-data__value" in html or "list__estimate-summary" in html:
        return parse_ebay_new(html, MALL_NEW, name)
    if "font-tahoma" in html:
        return parse_ebay_old(html, MALL_OLD, name)
    raise ValueError("지원하지 않는 견적서 양식입니다.")


# ---------------------------------------------------------------- 품명 다듬기 (규칙)

OPT_NO_RE = re.compile(r"^\s*\d{1,3}(?:-\d{1,3})?\)\s*")
SELECT_GRP_RE = re.compile(r"\(\s*[\d]+\s*~\s*[\d]+\s*\)\s*선택\s*/\s*")
LEAD_NO_RE = re.compile(r"^\s*\d{1,2}[.)]?\s+")
PACK_RE = re.compile(r"[xX]\s*(\d+)\s*(묶음|개입|세트|팩)")

LISTY_WORDS = ("모음", "모음전", "세트", "선택", "종류", "외", "인기", "베스트", "추천")


def looks_listy(name):
    words = name.split()
    hits = sum(1 for w in LISTY_WORDS if w in name)
    return len(words) >= 8 and hits >= 1


def strip_option_numbers(text):
    parts = [p.strip() for p in text.split(" + ")]
    parts = [OPT_NO_RE.sub("", p).strip() for p in parts]
    return " + ".join(p for p in parts if p)


def rule_refine(name, spec):
    name = (name or "").strip()
    spec = (spec or "").strip()

    had_select_group = bool(SELECT_GRP_RE.search(spec))
    spec = SELECT_GRP_RE.sub("", spec)
    spec = strip_option_numbers(spec)

    if spec and name:
        for sep in ("/", "+"):
            if spec.startswith(name):
                spec = spec[len(name):].lstrip(" " + sep).strip()
                break

    if spec and had_select_group and looks_listy(name):
        head = spec.split(" + ")[-1].strip() if " + " in spec else spec
        head = LEAD_NO_RE.sub("", head).strip()
        if head:
            m = PACK_RE.search(head)
            pack = ""
            if m:
                pack = f"{m.group(1)}{m.group(2)}"
                head = PACK_RE.sub("", head).strip()
            m2 = re.match(r"^([^(]{1,20}?)\s*\((.+)\)\s*$", head)
            if m2:
                new_name, inner = m2.group(1).strip(), m2.group(2).strip()
                new_spec = ", ".join(x for x in (inner, pack) if x)
            else:
                new_name, new_spec = head, pack
            if new_name:
                return new_name, new_spec

    return name, spec


# ---------------------------------------------------------------- 품명 다듬기 (Gemini, 브라우저 fetch)

GEMINI_MODEL = "gemini-2.5-flash"
THINKING_BUDGET = 0

PROMPT = """너는 학교 회계 담당자가 오픈마켓 견적서를 정리하는 일을 돕는다.

오픈마켓 상품명은 검색에 걸리려고 온갖 단어를 나열해 놓아서, 정작 무엇을 샀는지는
옵션에만 드러나는 경우가 많다. 아래 각 항목의 '원본품명'과 '원본규격'을 읽고
견적서에 쓸 '품명'과 '규격'으로 다시 나눠라.

규칙:
1. 품명은 실제로 구매한 물건 하나를 가리키는 간결한 이름으로 쓴다. (예: "철끈", "3M 테이프 522D")
2. 광고 문구와 나열된 딴 상품 이름은 버린다. (예: "인기 베스트", "학생용품 문구 볼펜 화일 가위")
3. 브랜드/모델명이 있으면 품명에 남긴다. (예: "모나미 네임펜M", "AULA F108Pro")
4. 색상·치수·수량단위·재질처럼 물건을 특정하는 값은 규격으로 보낸다. (예: "블랙", "0.4mm", "12자루")
5. 규격 항목은 쉼표로 구분한다.
6. **원본에 없는 정보를 지어내지 마라.** 숫자, 모델명, 색상, 규격을 바꾸거나 추측해서 채우지 마라.
   원본이 애매하면 애매한 대로 남겨라.
7. 원본 품명에 붙어 있는 띄어쓰기 오류는 자연스럽게 고쳐도 된다. (예: "톡톡파스텔" -> "톡톡 파스텔")
8. 품명이 이미 깔끔하면 거의 그대로 두어라.
9. **원본규격에 적힌 값은 하나도 빠뜨리지 마라.** 원본규격은 실제로 고른 옵션이라 무엇을
   샀는지 가리는 정보다. 품명에 넣지 않은 값은 반드시 규격에 남겨라.
   (예: 원본규격 "01 프락시누스" -> 규격에 "프락시누스"가 반드시 있어야 한다)
10. 원본품명에 섞인 수량·치수·용량 같은 '값'도 규격에 남겨라.
    (예: "100pcs 발사 나무 스틱" -> 규격에 "100pcs")
    단 이것이 규칙 2보다 우선하지는 않는다. 광고 문구나 용도 나열은 규격에도 넣지 말고 버려라.
    (예: "휴대용 디지털 풍속계 풍량계 바람측정기 캠핑 골프 아웃도어" 같은 수식어는 규격에도 쓰지 않는다)

입력은 JSON 배열이고, 각 원소는 {"i": 번호, "원본품명": "...", "원본규격": "..."} 이다.
출력도 같은 길이의 JSON 배열로, 각 원소는 {"i": 번호, "품명": "...", "규격": "..."} 이다.
설명 없이 JSON 배열만 출력하라.

입력:
"""

BATCH = 20
WORKERS = 4
MAX_RETRIES = 3   # 429(한도 초과)·503(과부하) 재시도 횟수 — 무료 티어 대응

PRICE = {
    "gemini-2.5-flash-lite": (0.10, 0.40),
    "gemini-2.5-flash": (0.30, 2.50),
}
USD_KRW = 1400


def estimate_won(tok_in, tok_out):
    pin, pout = PRICE.get(GEMINI_MODEL, (0.30, 2.50))
    usd = tok_in / 1_000_000 * pin + tok_out / 1_000_000 * pout
    return usd * USD_KRW


async def _retry_wait(resp, attempt):
    """429/503 재시도 대기 시간(초). 서버가 알려준 retryDelay 를 우선 쓰고,
    없으면 2·4·8초 지수 백오프. 최대 60초로 제한한다."""
    wait = 2 ** (attempt + 1)
    try:
        err = await resp.json()
        for d in (err.get("error") or {}).get("details") or []:
            rd = d.get("retryDelay")
            if isinstance(rd, str) and rd.endswith("s"):
                wait = max(wait, float(rd[:-1]))
    except Exception:
        pass
    return min(wait, 60)


async def gemini_refine(items, api_key, progress=None):
    """품명/규격만 Gemini REST 로 다듬는다. 실패하면 규칙 기반 결과를 유지한다.
    브라우저 fetch(pyodide.http.pyfetch)를 asyncio 로 동시에 날린다."""
    from pyodide.http import pyfetch

    uniq = {}
    for it in items:
        uniq.setdefault((it.raw_name, it.raw_spec), None)
    pairs = list(uniq.keys())
    if not pairs:
        return False, "정리할 품명이 없습니다."

    chunks = [pairs[s:s + BATCH] for s in range(0, len(pairs), BATCH)]
    url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
           f"{GEMINI_MODEL}:generateContent?key={api_key}")
    sem = asyncio.Semaphore(WORKERS)
    state = {"done": 0, "fail": 0, "rate_limited": 0, "retries": 0,
             "tok_in": 0, "tok_out": 0, "err_msg": ""}

    async def work(chunk):
        payload = [{"i": i, "원본품명": n, "원본규격": sp}
                   for i, (n, sp) in enumerate(chunk)]
        body = {
            "contents": [{"parts": [{"text": PROMPT + json.dumps(
                payload, ensure_ascii=False, indent=1)}]}],
            "generationConfig": {
                "temperature": 0,
                "responseMimeType": "application/json",
                "thinkingConfig": {"thinkingBudget": THINKING_BUDGET},
            },
        }
        data = None
        rate_limited = False
        async with sem:
            for attempt in range(MAX_RETRIES + 1):
                try:
                    resp = await pyfetch(
                        url, method="POST",
                        headers={"Content-Type": "application/json"},
                        body=json.dumps(body),
                    )
                except Exception as e:
                    state["err_msg"] = state["err_msg"] or f"네트워크 오류: {e}"
                    break

                if resp.status == 200:
                    try:
                        res = await resp.json()
                        text = res["candidates"][0]["content"]["parts"][0]["text"]
                        data = json.loads(text)
                        u = res.get("usageMetadata") or {}
                        state["tok_in"] += u.get("promptTokenCount", 0) or 0
                        state["tok_out"] += (u.get("candidatesTokenCount", 0) or 0) + (
                            u.get("thoughtsTokenCount", 0) or 0)
                    except Exception as e:
                        state["err_msg"] = state["err_msg"] or f"응답 해석 실패: {e}"
                    break

                # 429(요청 한도 초과)·503(과부하)은 잠시 쉬었다 재시도한다.
                # 유료 키는 여기 거의 안 걸리므로 풀 속도 그대로,
                # 무료 키는 자동으로 느려지며 끝까지 완주한다.
                if resp.status in (429, 503):
                    if attempt < MAX_RETRIES:
                        state["retries"] += 1
                        wait = await _retry_wait(resp, attempt)
                        if progress:
                            progress(state["done"], len(pairs),
                                     f"무료 티어 한도로 {wait:.0f}초 대기 후 재시도 중… "
                                     f"({state['done']}/{len(pairs)})")
                        await asyncio.sleep(wait)
                        continue
                    rate_limited = True
                    state["err_msg"] = state["err_msg"] or (
                        f"요청 한도 초과(HTTP {resp.status})")
                    break

                # 그 외 HTTP 오류: 서버가 준 메시지를 그대로 남겨 원인 파악에 쓴다.
                try:
                    err = await resp.json()
                    state["err_msg"] = state["err_msg"] or (
                        (err.get("error") or {}).get("message")
                        or f"HTTP {resp.status}")
                except Exception:
                    state["err_msg"] = state["err_msg"] or f"HTTP {resp.status}"
                break

        if data is None:
            state["fail"] += len(chunk)
            if rate_limited:
                state["rate_limited"] += len(chunk)
            if progress:
                progress(state["done"], len(pairs))
            return

        got = 0
        for row in data:
            try:
                i = int(row["i"])
            except (KeyError, TypeError, ValueError):
                continue
            if 0 <= i < len(chunk):
                name = (row.get("품명") or "").strip()
                spec = (row.get("규격") or "").strip()
                if name:
                    uniq[chunk[i]] = (name, spec)
                    got += 1
        state["done"] += got
        if progress:
            progress(state["done"], len(pairs))

    await asyncio.gather(*[work(c) for c in chunks])

    if not state["done"]:
        if state["rate_limited"] or state["retries"]:
            return False, (
                "무료 티어 요청 한도(분당·하루)를 초과해 Gemini 정리를 못 했습니다. "
                "잠시 후 다시 시도하거나, 결제가 연결된 API 키를 쓰면 한도가 크게 늘어납니다. "
                "지금은 규칙 기반 결과만 적용했습니다.")
        if state["err_msg"]:
            return False, (
                f"Gemini 응답 실패: {state['err_msg']} — 규칙 기반 결과만 적용했습니다.")
        return False, "Gemini 응답을 받지 못해 규칙 기반 결과만 적용했습니다."

    for it in items:
        got = uniq.get((it.raw_name, it.raw_spec))
        if got:
            it.name, it.spec = got

    msg = f"Gemini({GEMINI_MODEL})로 품명 {state['done']}종을 정리했습니다. (고유 상품 {len(pairs)}종)"
    if state["fail"]:
        other = state["fail"] - state["rate_limited"]
        if state["rate_limited"]:
            msg += f" — {state['rate_limited']}종은 무료 티어 요청 한도 초과"
            if other > 0:
                msg += f", {other}종은 응답 실패"
            msg += "로 원문 그대로 두었습니다. (잠시 후 재시도하거나 결제 연결된 키 권장)"
        else:
            msg += f" — {state['fail']}종은 응답을 못 받아 원문 그대로 두었습니다."
            if state["err_msg"]:
                msg += f" ({state['err_msg']})"
    if state["tok_in"]:
        won = estimate_won(state["tok_in"], state["tok_out"])
        msg += (f" 토큰 입력 {state['tok_in']:,} / 출력 {state['tok_out']:,}"
                f" — 이번 호출 약 {won:.1f}원")
    if state["retries"] and not state["rate_limited"]:
        msg += (" ※ 무료 티어 한도로 재시도하며 처리해 다소 느렸습니다. "
                "결제가 연결된 API 키를 쓰면 훨씬 빠릅니다.")
    return True, msg


async def refine_all(sheets, api_key="", use_gemini=True, progress=None):
    items = [it for s in sheets for it in s.items]
    for it in items:
        it.keep_raw()
        it.name, it.spec = rule_refine(it.raw_name, it.raw_spec)
    if not use_gemini or not api_key:
        return False, "규칙 기반으로만 정리했습니다. (Gemini 미사용)"
    return await gemini_refine(items, api_key, progress=progress)


# ---------------------------------------------------------------- 표 만들기

HEADER = ["품명", "규격", "단가", "수량", "금액", "비고"]


def unit_price(amount, qty):
    if not qty:
        return 0
    u = amount / qty
    return int(u) if float(u).is_integer() else round(u, 2)


def order_sheets(sheets):
    return sorted(
        sheets,
        key=lambda s: MALL_ORDER.index(s.mall) if s.mall in MALL_ORDER else len(MALL_ORDER),
    )


def is_shipping_row(row):
    return str(row[0]).endswith("배송비")


def build_rows(sheets):
    rows = []
    for s in order_sheets(sheets):
        for it in s.items:
            rows.append([it.name, it.spec, unit_price(it.amount, it.qty),
                         it.qty, it.amount, it.note])
        if s.shipping:
            rows.append([f"{s.mall} 배송비", "", s.shipping, 1, s.shipping, ""])
    return rows


def to_tsv(rows):
    lines = ["\t".join(HEADER)]
    for r in rows:
        lines.append("\t".join(str(c) for c in r))
    return "\n".join(lines)


def build_report(sheets, rows):
    out = []
    grand = 0
    for s in order_sheets(sheets):
        goods = sum(i.amount for i in s.items)
        total = goods + s.shipping
        grand += total
        ok = (s.stated_total == 0) or (total == s.stated_total)
        mark = "일치" if ok else f"불일치! 견적서상 {s.stated_total:,}원"
        out.append(
            f"[{s.mall}] {s.source}\n"
            f"    항목 {len(s.items)}건  상품 {goods:,}원 + 배송비 {s.shipping:,}원 "
            f"= {total:,}원  ({mark})"
        )
    out.append("")
    out.append(f"총 {len(rows)}행 (배송비 포함), 합계 {grand:,}원")

    odd = [r for r in rows if not float(r[2]).is_integer()]
    if odd:
        out.append("")
        out.append(f"※ 단가가 딱 나누어떨어지지 않는 항목 {len(odd)}건 (금액 기준은 정확함):")
        for r in odd[:5]:
            out.append(f"    - {str(r[0])[:30]} : {r[4]:,}원 / {r[3]}개 = {r[2]}")
    return "\n".join(out)


# ---------------------------------------------------------------- 엑셀 보고서 (bytes 반환)

def write_report_xlsx(sheets, rows, refine_msg=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(name="맑은 고딕", size=14, bold=True)
    head_font = Font(name="맑은 고딕", size=10, bold=True)
    body_font = Font(name="맑은 고딕", size=10)
    money = "#,##0"

    wb = Workbook()

    ws = wb.active
    ws.title = "견적서"
    ws["A1"] = "견 적 서"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"작성일 {datetime.date.today():%Y-%m-%d}"
    ws["A2"].font = Font(name="맑은 고딕", size=9, color="808080")
    ws.merge_cells("A2:F2")

    r = 4
    for c, h in enumerate(HEADER, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = box
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ship_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in rows:
        r += 1
        ship = is_shipping_row(row)
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name="맑은 고딕", size=10, bold=True) if ship else body_font
            cell.border = box
            if ship:
                cell.fill = ship_fill
            if c in (3, 4, 5):
                cell.number_format = money
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(c <= 2))

    first, last = 5, r
    r += 1
    ws.cell(row=r, column=1, value="합계").font = head_font
    ws.cell(row=r, column=1).fill = head_fill
    total_cell = ws.cell(row=r, column=5, value=f"=SUM(E{first}:E{last})")
    total_cell.font = head_font
    total_cell.fill = head_fill
    total_cell.number_format = money
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = box
        if c != 1 and c != 5:
            ws.cell(row=r, column=c).fill = head_fill

    for col, w in zip("ABCDEF", (44, 30, 11, 7, 13, 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{last}"

    # ---------- 시트 2: 요약
    s2 = wb.create_sheet("요약")
    s2["A1"] = "정리 요약"
    s2["A1"].font = title_font
    s2.merge_cells("A1:F1")

    hdr2 = ["몰", "원본 파일", "항목수", "상품금액", "배송비", "합계", "견적서 총액", "검증"]
    for c, h in enumerate(hdr2, start=1):
        cell = s2.cell(row=3, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = box
        cell.alignment = Alignment(horizontal="center")

    r2 = 3
    for s in order_sheets(sheets):
        r2 += 1
        goods = sum(i.amount for i in s.items)
        total = goods + s.shipping
        ok = (s.stated_total == 0) or (total == s.stated_total)
        vals = [s.mall, s.source, len(s.items), goods, s.shipping, total,
                s.stated_total or "", "일치" if ok else "불일치"]
        for c, v in enumerate(vals, start=1):
            cell = s2.cell(row=r2, column=c, value=v)
            cell.font = body_font
            cell.border = box
            if c in (4, 5, 6, 7):
                cell.number_format = money
        vcell = s2.cell(row=r2, column=8)
        vcell.font = Font(name="맑은 고딕", size=10, bold=True,
                          color="1F7A1F" if ok else "C00000")
        vcell.alignment = Alignment(horizontal="center")

    r2 += 1
    s2.cell(row=r2, column=1, value="총계").font = head_font
    s2.cell(row=r2, column=1).fill = head_fill
    for c in (3, 4, 5, 6):
        col = get_column_letter(c)
        cell = s2.cell(row=r2, column=c, value=f"=SUM({col}4:{col}{r2 - 1})")
        cell.font = head_font
        cell.fill = head_fill
        cell.border = box
        if c != 3:
            cell.number_format = money
    for c in (1, 2, 7, 8):
        s2.cell(row=r2, column=c).fill = head_fill
        s2.cell(row=r2, column=c).border = box

    r2 += 2
    s2.cell(row=r2, column=1, value="비고").font = head_font
    notes = [
        refine_msg,
        "금액은 각 견적서의 실제 결제금액(할인 반영)을 그대로 옮긴 값입니다.",
        "단가 = 금액 ÷ 수량. 나누어떨어지지 않으면 소수로 두어 금액 합계를 정확히 맞췄습니다.",
        "배송비는 몰별로 분리해 별도 항목으로 넣었습니다.",
        "품명/규격은 읽기 쉽게 재작성한 것입니다. 원문은 '원문대조' 시트에 있습니다.",
    ]
    for n in notes:
        if not n:
            continue
        r2 += 1
        c = s2.cell(row=r2, column=1, value="· " + n)
        c.font = Font(name="맑은 고딕", size=9)
        s2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)

    for col, w in zip("ABCDEFGH", (12, 26, 8, 14, 11, 14, 14, 9)):
        s2.column_dimensions[col].width = w

    # ---------- 시트 3: 원문대조
    s3 = wb.create_sheet("원문대조")
    hdr3 = ["몰", "품명(정리)", "규격(정리)", "원본 상품명", "원본 옵션", "수량", "금액"]
    for c, h in enumerate(hdr3, start=1):
        cell = s3.cell(row=1, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = box
    r3 = 1
    for s in order_sheets(sheets):
        for it in s.items:
            r3 += 1
            vals = [it.mall, it.name, it.spec, it.raw_name, it.raw_spec, it.qty, it.amount]
            for c, v in enumerate(vals, start=1):
                cell = s3.cell(row=r3, column=c, value=v)
                cell.font = body_font
                cell.border = box
                if c in (6, 7):
                    cell.number_format = money
    for col, w in zip("ABCDEFG", (10, 34, 26, 52, 34, 7, 12)):
        s3.column_dimensions[col].width = w
    s3.freeze_panes = "A2"
    s3.auto_filter.ref = f"A1:G{r3}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 에듀파인 업로드 서식
# 에듀파인 '물품내역' 업로드 서식에 그대로 맞춘 파일을 만든다. 내려받아 바로 업로드할 수 있다.
# (원본 서식은 바이너리 .xls 지만, 실측 결과 .xlsx 도 에듀파인이 그대로 인식·업로드된다.
#  그래서 별도 라이브러리 없이 openpyxl 로 .xlsx 를 만든다.)
# 인식 조건상 시트명은 반드시 '물품내역' 이어야 하고, 이 한 장만 담는다(다른 시트 없음).
#   열: 품명 · 규격 · 수량 · 단위 · 예상단가 · 예상금액 · 용도
#   단위는 '개', 용도는 빈칸. 배송비도 한 행으로 들어간다(서식 원본과 동일).

EDUFINE_HEADER = ["품명", "규격", "수량", "단위", "예상단가", "예상금액", "용도"]


def write_edufine_xlsx(rows):
    """build_rows() 결과를 에듀파인 물품내역 서식(.xlsx, 시트명 '물품내역')으로 만든다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "물품내역"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    money = "#,##0"

    for c, h in enumerate(EDUFINE_HEADER, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.alignment = center

    r = 1
    for row in rows:
        r += 1
        name, spec, unitp, qty, amount, _note = row
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=spec)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value="개")          # 단위 (서식이 전부 '개')
        ws.cell(row=r, column=5, value=unitp).number_format = money   # 예상단가 = 금액 ÷ 수량
        ws.cell(row=r, column=6, value=amount).number_format = money  # 예상금액 = 실결제(정확)
        ws.cell(row=r, column=7, value="")             # 용도 (빈칸)

    for col, w in zip("ABCDEFG", (44, 30, 7, 6, 11, 13, 16)):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 진입점 (JS 가 호출)

async def convert(files, api_key="", use_gemini=True, progress=None):
    """files: [(파일명, bytes), ...] -> 결과 dict.
    반환: {ok, tsv, report, refine_msg, xlsx(bytes|None), summary[], errors[]}"""
    sheets, errors = [], []
    for name, data in files:
        try:
            sheets.append(parse_file(name, bytes(data)))
        except Exception as e:
            errors.append(f"{name} : {e}")

    if not sheets:
        return {
            "ok": False, "tsv": "", "report": "",
            "refine_msg": "", "xlsx": None, "edufine": None,
            "summary": [], "errors": errors,
        }

    _, refine_msg = await refine_all(
        sheets, api_key=api_key, use_gemini=use_gemini, progress=progress)

    rows = build_rows(sheets)
    tsv = to_tsv(rows)
    report = build_report(sheets, rows)
    try:
        xlsx = write_report_xlsx(sheets, rows, refine_msg)
    except Exception as e:
        xlsx = None
        errors.append(f"엑셀 생성 실패: {e}")
    try:
        edufine = write_edufine_xlsx(rows)
    except Exception as e:
        edufine = None
        errors.append(f"에듀파인 서식 생성 실패: {e}")

    summary = []
    for s in order_sheets(sheets):
        goods = sum(i.amount for i in s.items)
        total = goods + s.shipping
        ok = (s.stated_total == 0) or (total == s.stated_total)
        summary.append({
            "mall": s.mall, "source": s.source, "items": len(s.items),
            "goods": goods, "shipping": s.shipping, "total": total,
            "stated": s.stated_total, "ok": ok,
        })

    return {
        "ok": True, "tsv": tsv, "report": report,
        "refine_msg": refine_msg, "xlsx": xlsx, "edufine": edufine,
        "summary": summary, "errors": errors,
    }
